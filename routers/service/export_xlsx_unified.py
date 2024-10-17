from io import BytesIO

import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from starlette.responses import StreamingResponse

from .export_xlsx import get_output


def export_record_unified(response):
    output = get_output(response)
    filename = 'Planilha'

    if '.' in response['filename']:
        filename = response['filename'].split('.')[0]

    unified_workbook = openpyxl.Workbook()
    unified_worksheet = unified_workbook.active
    unified_worksheet.sheet_view.showGridLines = False
    unified_worksheet.title = filename

    output.seek(0)
    workbook = openpyxl.load_workbook(output)

    row_offset = 1
    for index, sheet_name in enumerate(workbook.sheetnames):

        sheet = workbook[sheet_name]

        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        bold_font = Font(bold=True)

        alignment = Alignment(horizontal='center', vertical='center')

        if index == 0:
            cell_title = unified_worksheet.cell(row=row_offset, column=1, value=f"{filename.upper()}")

            cell_title.border = border
            cell_title.font = bold_font
            cell_title.alignment = alignment

            row_offset +=2

        cell = unified_worksheet.cell(row=row_offset, column=1, value=f"{sheet_name}")
        cell.border = border
        cell.font = bold_font
        cell.alignment = alignment

        for index_row, row in enumerate(sheet.iter_rows()):

            for cell in row:
                new_cell = unified_worksheet.cell(row=cell.row + row_offset, column=cell.col_idx, value=cell.value)
                __copy_cell_style(cell, new_cell)
                if index_row == 0:
                    new_cell.alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')

                column_letter = openpyxl.utils.get_column_letter(cell.col_idx)

                current_width = unified_worksheet.column_dimensions[column_letter].width or 0
                new_width = max(current_width, len(str(cell.value)) + 2)
                unified_worksheet.column_dimensions[column_letter].width = new_width

            unified_worksheet.row_dimensions[cell.row + row_offset].height = sheet.row_dimensions[cell.row].height

            if index_row == 0:
                height = unified_worksheet.row_dimensions[row_offset+1].height
                unified_worksheet.row_dimensions[row_offset+1].height =  30 if height is None else height * 2


        row_offset += sheet.max_row
        row_offset += 2
    
    output = BytesIO()
    unified_workbook.save(output)
    output.seek(0)

    return StreamingResponse(output,
                             media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={"Content-Disposition": f"attachment; filename={'filename.xlsx'}"})

def __copy_cell_style(source_cell, target_cell):
    if source_cell.font:
        target_cell.font = openpyxl.styles.Font(
            name=source_cell.font.name,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
    if source_cell.fill:
        target_cell.fill = openpyxl.styles.PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
    if source_cell.border:
        target_cell.border = openpyxl.styles.Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom
        )
    if source_cell.alignment:
        target_cell.alignment = openpyxl.styles.Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )


